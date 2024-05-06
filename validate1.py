import time
import csv
import akshare as ak
import pandas as pd
import os
import datetime
import pytz
import utils
import threading
import concurrent.futures
from datetime import datetime as dt
from openpyxl import load_workbook
from openpyxl import Workbook


# 对三连跌+十字星+反包的验证
def getTimeInterval():
    timelist = utils.gettardeDayList('20220101', '20240401')
    return timelist


# 1. 获取出所有符合条件的股票以及相应的时间段
def getMatchStocks():
    timelist = getTimeInterval()

    results = []
    futurelist = []
    allstocks = utils.getAllStocks()

    tmpstocks = allstocks[0, 10]

    # 线程池
    with concurrent.futures.ThreadPoolExecutor(max_workers=200) as executor:
        for row in tmpstocks.itertuples():
            code = getattr(row, '证券代码')
            name = getattr(row, '证券简称')

            future = executor.submit(getMatchStocks_thread, code, name, timelist)
            futurelist.append(future)

        for future in concurrent.futures.as_completed(futurelist):
            result = future.result()
            if result != None:
                results.append(result)

    return results


def getMatchStocks_thread(code, name, timelist):
    print(code, name)
    checkList = []

    ids = 0
    list_length = len(timelist)
    while ids < list_length - 5:
        start_date = timelist[ids]
        # print(start_date)
        end_date = timelist[ids + 4]

        try:
            # 取5天数据
            stock_zh_a_hist_df = ak.stock_zh_a_hist(
                symbol=code,
                start_date=start_date,
                end_date=end_date
            )

            zhangdiefu = [0, 0, 0, 0, 0]
            kaipanjia = [0, 0, 0, 0, 0]
            shoupanjia = [0, 0, 0, 0, 0]
            max_v = [0, 0, 0, 0, 0]
            min_v = [0, 0, 0, 0, 0]

            i = 0
            for item in stock_zh_a_hist_df.itertuples():
                zhangdiefu[i] = getattr(item, '涨跌幅')
                kaipanjia[i] = getattr(item, '开盘')
                shoupanjia[i] = getattr(item, '收盘')
                max_v[i] = getattr(item, '最高')
                min_v[i] = getattr(item, '最低')
                i = i + 1

            # 1. 前三天下跌
            if zhangdiefu[0] < 0 and zhangdiefu[1] < 0 and zhangdiefu[2] < 0:
                # 2. 第四天十字星
                body_length = abs(kaipanjia[3] - shoupanjia[3])
                upper_shadow = max_v[3] - max(kaipanjia[3], shoupanjia[3])
                lower_shadow = min(kaipanjia[3], shoupanjia[3]) - min_v[3]
                if body_length < 0.05 * shoupanjia[3] and upper_shadow > 2 * body_length and lower_shadow > 2 * body_length:
                    # 3. 第五天大阳线
                    if zhangdiefu[4] > 6:
                        print('符合条件的数据为', code, name, start_date)
                        checkList.append((code, name, start_date))
        except Exception as e:
            continue
        finally:
            ids = ids + 1
    return checkList


# 2. 将数据写入文件
def writefile(result):
    with open('example.csv', 'w', newline='') as file:
        writer = csv.writer(file)
        writer.writerows(result)


# 3. 手动处理文件后，从文件中进行计算
# 符合条件股票，在之后1-5个交易日中的涨跌幅
def calcRate(filename):
    item_list = []
    # 打开.xlsx文件
    workbook = load_workbook(filename)

    # 选择工作表
    sheet = workbook.active

    # 读取数据
    for row in sheet.iter_rows(values_only=True):
        item_list.append(row)

    # 关闭工作簿
    workbook.close()

    item_list_act = item_list[1:]


    tradelist = utils.gettardeDayList('20210101', '20240401')

    new_item_list = []
    for item_tup in item_list_act:
        code = item_tup[0].strip()
        name = item_tup[1]
        start_date = item_tup[2]
        result = calc_stock(code, name, start_date, tradelist)
        print(result)
        if result != None:
            new_item_list.append(result)

    return new_item_list


def calc_stock(code, name, start_date, tradelist):
    try:
        start_date_index = tradelist.index(str(start_date))
        new_start_date = tradelist[start_date_index + 5]
        end_date = tradelist[start_date_index + 9]

        # 获取五天的历史数据
        stock_zh_a_hist_df = ak.stock_zh_a_hist(
            symbol=code,
            start_date=new_start_date,
            end_date=end_date
        )
        print(stock_zh_a_hist_df)

        zhangdiefu = [0, 0, 0, 0, 0]
        i = 0
        for item in stock_zh_a_hist_df.itertuples():
            zhangdiefu[i] = getattr(item, '涨跌幅')
            i = i + 1
        return (code, name, start_date, zhangdiefu[0], zhangdiefu[1], zhangdiefu[2], zhangdiefu[3], zhangdiefu[4])

    except Exception as e:

        return None

def writefiles(new_item_list):
    # 创建一个工作簿对象
    workbook = Workbook()

    # 激活默认的工作表
    sheet = workbook.active

    # 在工作表中写入数据
    sheet['A1'] = 'code'
    sheet['B1'] = 'name'
    sheet['C1'] = '开始日期'
    sheet['D1'] = '第一天'
    sheet['E1'] = '第二天'
    sheet['F1'] = '第三天'
    sheet['G1'] = '第四天'
    sheet['H1'] = '第五天'


    data = new_item_list

    for row in data:
        sheet.append(row)

    # 保存工作簿到文件
    workbook.save('output.xlsx')


if __name__ == '__main__':
    new_item_list = calcRate('validate.xlsx')
    # writefiles(new_item_list)
